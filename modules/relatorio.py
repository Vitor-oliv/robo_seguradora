'''
Módulo de Relatórios Excel (Diário e Completo)
===============================================
'''

import os
from datetime import datetime
import logging
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

class GeradorRelatorio:
    "Classe para gerar relatórios Excel diários."
    
    def __init__(self, config):
        self.config = config
        os.makedirs(config.RELATORIO_DIR, exist_ok=True)
    
    def gerar_excel_estoque(self, pecas_processadas):
        "Gera ou atualiza um arquivo Excel diário com todos os itens consultados."
        try:
            if not pecas_processadas:
                logger.warning("Nenhuma peça para registrar no relatório.")
                return None
            
            # Nome do arquivo baseado na data atual para criar um arquivo por dia
            data_hoje = datetime.now().strftime("%Y-%m-%d")
            filename = f"relatorio_diario_{data_hoje}.xlsx"
            filepath = os.path.join(self.config.RELATORIO_DIR, filename)
            
            # Verifica se o arquivo já existe para decidir se cria um novo ou anexa
            if os.path.exists(filepath):
                wb = load_workbook(filepath)
                ws = wb.active
                logger.info(f"Anexando dados ao relatório existente: {filepath}")
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Conferência Diária"
                # Cabeçalho apenas para arquivos novos
                ws.append(["DATA/HORA", "PARTNUMBER", "STATUS", "MARCADO NO SITE"])
                logger.info(f"Criando novo relatório diário: {filepath}")

            # Adiciona os dados de todas as peças processadas
            for p in pecas_processadas:
                status_texto = "DISPONÍVEL" if p.tem_estoque else "INDISPONÍVEL"
                marcado_texto = "SIM" if p.tem_estoque else "NÃO"
                
                ws.append([
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    p.partnumber,
                    status_texto,
                    marcado_texto
                ])
            
            wb.save(filepath)
            logger.info(f"✓ Relatório atualizado com sucesso: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"✗ Erro ao gerar relatório Excel: {e}")
            return None
