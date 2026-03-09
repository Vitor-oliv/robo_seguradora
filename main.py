import os
import logging
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import config
from modules.estoque import ConsultorEstoque

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

robo_instancia = None

class RoboSeguradora:
    def __init__(self):
        self.driver = None
        self.consultor = None

    def setup_driver(self):
        chrome_options = Options()
        chrome_options.add_argument(f"--user-data-dir={config.PERFIL_CHROME_DIR}")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        chrome_options.add_experimental_option("useAutomationExtension", False)
        if config.HEADLESS_MODE: chrome_options.add_argument("--headless")
        
        try:
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.driver.maximize_window()
            self.driver.get(config.SEGURADORA_URL)
            return True
        except Exception as e:
            logger.error(f"Erro ao abrir Chrome: {e}")
            return False

    def fazer_login(self):
        try:
            wait = WebDriverWait(self.driver, config.TIMEOUT_PADRAO)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, config.SELETORES["campo_empresa"]))).send_keys(config.SEGURADORA_EMPRESA)
            self.driver.find_element(By.CSS_SELECTOR, config.SELETORES["campo_login"]).send_keys(config.SEGURADORA_LOGIN)
            self.driver.find_element(By.CSS_SELECTOR, config.SELETORES["campo_senha"]).send_keys(config.SEGURADORA_SENHA)
            self.driver.find_element(By.CSS_SELECTOR, config.SELETORES["botao_entrar"]).click()
            time.sleep(config.DELAY_ENTRE_ACOES)
            return True
        except Exception as e:
            logger.error(f"Finalizado")
            return True

    def obter_marca_e_estado(self):
        try:
            marca_element = self.driver.find_element(By.CSS_SELECTOR, config.SELETORES["marca"])
            marca_text = marca_element.text.strip().upper()
            marca = ""
            for key in config.TABELA_DESCONTOS.keys():
                if key in marca_text:
                    marca = key
                    break
            
            estado_element = self.driver.find_element(By.CSS_SELECTOR, config.SELETORES["estado"])
            estado_text = estado_element.text.strip().upper()
            estado = "SP" if "SP" in estado_text else "OUTROS"
            
            logger.info(f"Marca detectada: {marca}, Estado detectado: {estado}")
            return marca, estado
        except Exception as e:
            logger.warning(f"Não foi possível obter marca ou estado: {e}")
            return "", "OUTROS"

    def processar_cotacao_e_aplicar_desconto(self):
        if not self.driver:
            return {"erro": "Navegador não está aberto!"}
            
        resultados = []
        total_pecas = 0
        em_estoque = 0
        
        marca, estado = self.obter_marca_e_estado()
        desconto_percentual = 0

        if marca and marca in config.TABELA_DESCONTOS:
            desconto_percentual = config.TABELA_DESCONTOS[marca].get(estado, config.TABELA_DESCONTOS[marca]["OUTROS"])
            logger.info(f"Desconto calculado para {marca} em {estado}: {desconto_percentual}%")
        else:
            logger.warning(f"Marca {marca} não encontrada na tabela de descontos. Aplicando 0% de desconto.")

        i = 0
        while True:
            try:
                id_part = f"{config.SELETORES['id_partnumber_prefix']}{i}"
                id_check = f"{config.SELETORES['id_checkbox_prefix']}{i}"
                
                elementos = self.driver.find_elements(By.ID, id_part)
                if not elementos:
                    break
                
                part_number = elementos[0].get_attribute("value")
                if not part_number:
                    break
                
                total_pecas += 1
                
                tem_estoque = self.consultar_estoque_real(part_number)
                
                if tem_estoque:
                    em_estoque += 1
                    checkbox = self.driver.find_element(By.ID, id_check)
                    if not checkbox.is_selected():
                        checkbox.click()
                    
                    # Aplica o desconto se houver e o item estiver em estoque
                    if desconto_percentual > 0:
                        # FORMATAÇÃO CORRIGIDA: Usa vírgula e duas casas decimais (ex: 52,00)
                        desconto_formatado = f"{float(desconto_percentual):.2f}".replace(".", ",")
                        
                        campo_desconto = self.driver.find_element(By.CSS_SELECTOR, config.SELETORES["campo_desconto"])
                        campo_desconto.clear()
                        campo_desconto.send_keys(desconto_formatado)
                        time.sleep(config.DELAY_ENTRE_ACOES)
                        
                        botao_aplicar = self.driver.find_element(By.CSS_SELECTOR, config.SELETORES["botao_aplicar_desconto"])
                        botao_aplicar.click()
                        logger.info(f"Desconto de {desconto_formatado}% aplicado para o item {part_number}.")
                        time.sleep(config.DELAY_ENTRE_ACOES)

                resultados.append({
                    "part_number": part_number,
                    "estoque": "Sim" if tem_estoque else "Não",
                    "desconto_aplicado": f"{desconto_percentual}%" if tem_estoque and desconto_percentual > 0 else "Nenhum",
                    "data_hora": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                })
                i += 1
            except Exception as e:
                logger.error(f"Erro ao processar item {i}: {e}")
                break

        if resultados:
            self.salvar_relatorio(resultados)
            return {"sucesso": True, "mensagem": f"Processo concluído. Total de peças: {total_pecas}, Em estoque (marcados e com desconto): {em_estoque}."}
        else:
            return {"erro": "Nenhum item encontrado na tela para processar."}

    def consultar_estoque_real(self, part_number):
        if self.consultor is None:
            self.consultor = ConsultorEstoque()
        return self.consultor.verificar_disponibilidade(part_number)

    def salvar_relatorio(self, resultados):
        from openpyxl import Workbook, load_workbook
        data_hoje = datetime.now().strftime("%Y-%m-%d")
        filepath = os.path.join(config.RELATORIO_DIR, f"relatorio_{data_hoje}.xlsx")
        
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["DATA/HORA", "PARTNUMBER", "EM ESTOQUE", "DESCONTO APLICADO"])
        
        for res in resultados:
            ws.append([res["data_hora"], res["part_number"], res["estoque"], res["desconto_aplicado"]])
        wb.save(filepath)

def presence_of_element_with_id_or_selector(selector):
    def _predicate(driver):
        try:
            if selector.startswith("#"):
                return driver.find_element(By.ID, selector[1:])
            return driver.find_element(By.CSS_SELECTOR, selector)
        except:
            return False
    return _predicate

def iniciar_navegador():
    global robo_instancia
    if robo_instancia is None:
        robo_instancia = RoboSeguradora()
    
    if robo_instancia.setup_driver():
        robo_instancia.consultor = ConsultorEstoque()
        robo_instancia.fazer_login()
        return robo_instancia.driver
    return None

def conferir_estoque_e_aplicar_desconto():
    global robo_instancia
    if robo_instancia:
        return robo_instancia.processar_cotacao_e_aplicar_desconto()
    return {"erro": "O robô não foi iniciado corretamente."}

def fechar_tudo():
    global robo_instancia
    if robo_instancia and robo_instancia.driver:
        try:
            robo_instancia.driver.quit()
        except:
            pass
        robo_instancia = None
